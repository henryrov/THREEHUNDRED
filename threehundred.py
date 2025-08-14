import openpyxl
import datetime
import zoneinfo
import icalendar
import sys
import os

class MeetingPattern:
    def __init__(self, start, end, until, location, weekdays):
        self.start_time = start
        self.end_time = end
        self.until_date = until
        self.location = location
        self.weekdays = weekdays

    def print_patterns(self):
        print("Class starts on:", self.start_time)
        print("Class ends at:", self.end_time)
        print("Last class is on:", self.until_date)
        print("Weekdays:", self.weekdays)


class CourseSection:
    def __init__(self, name, meeting_patterns):
        self.name = name;
        self.meeting_patterns = meeting_patterns;

    def print_info(self):
        print("Course section:", self.name)
        for i in range (0, len(self.meeting_patterns)):
            self.meeting_patterns[i].print_patterns()

        print()


def parse_section_name(name):
    slice_end = str(name).find(" - ")
    if slice_end < 0:
        print("Error: Section name not recognized")
        return 0;

    section_name = name[0:slice_end]
    return section_name


def parse_section_meeting_patterns(mp_string):
    patterns = []

    lines = str(mp_string).split("\n")
    for line in range(0, len(lines)):
        substrings = lines[line].split("|")
        if len(substrings) == 1:
            # Empty line
            continue

        date_separator = substrings[0].index("- ") - 1

        start_date = datetime.datetime.fromisoformat(
            substrings[0][0:date_separator]
        )
        end_date = datetime.datetime.fromisoformat(
            substrings[0][0:date_separator]
        )
        until_date = datetime.datetime.fromisoformat(
            substrings[0][date_separator + 3:len(substrings[0]) - 1])

        weekdays = []
        days = substrings[1].lstrip().rstrip().split(" ")
        for day in days:
            weekdays.append(day[0:2].lower())

        times = substrings[2].split(" - ")
        start_hours = int(times[0][1:times[0].index(":")])
        if times[0].find("p.m.") >= 0 and start_hours != 12:
            start_hours += 12
        start_minutes = int(
            times[0][times[0].index(":") + 1:times[0].index(":") + 3]
        )

        end_hours = int(times[1][0:times[1].index(":")])
        if times[1].find("p.m.") >= 0 and end_hours != 12:
            end_hours += 12
        end_minutes = int(
            times[1][times[1].index(":") + 1:times[1].index(":") + 3]
        )

        start_date = start_date.replace(
            hour = start_hours, minute = start_minutes,
            tzinfo=zoneinfo.ZoneInfo("America/Vancouver")
        )
        end_date = end_date.replace(
            hour = end_hours, minute = end_minutes,
            tzinfo=zoneinfo.ZoneInfo("America/Vancouver")
        )
        until_date = until_date.replace(
            hour = end_hours, minute = end_minutes,
            tzinfo=zoneinfo.ZoneInfo("America/Vancouver")
        )

        location = substrings[3].lstrip().rstrip()

        meeting_pattern = MeetingPattern(
            start_date,
            end_date,
            until_date,
            location,
            weekdays
        )
        patterns.append(meeting_pattern)

    return patterns


def parse_sections(sheet, sections):
    # Data starts in Row 4

    for row in range(4, sheet.max_row + 1):
        # Sections names are in column 4 (D)

        cell = sheet.cell(row = row, column = 4)
        section_name = parse_section_name(cell.value)
        if section_name == 0:
            continue;

        # Meeting patterns are in column 10 (J)

        cell = sheet.cell(row = row, column = 10)
        mp_string = cell.value
        meeting_patterns = parse_section_meeting_patterns(mp_string)

        section = CourseSection(section_name, meeting_patterns)
        sections.append(section)

def gen_ics(sections):
    cal = icalendar.Calendar()

    cal.add('prodid', '-//henry.rov//THREEHUNDRED//EN')
    cal.add('version', '2.0')

    for section in sections:
        for pattern in section.meeting_patterns:
            event = icalendar.Event()
            event.add("summary", section.name);
            event.add("dtstart", pattern.start_time)
            event.add("dtstamp", datetime.datetime.utcnow())
            event.add("dtend", pattern.end_time)
            rrule = icalendar.vRecur(
                freq = "WEEKLY",
                until = pattern.until_date,
                byday = pattern.weekdays
            )
            event.add("rrule", rrule)
            event.add("location", pattern.location)
            cal.add_component(event)

    cal_file = open("out.ics", "wb")
    cal_file.write(cal.to_ical())
    cal_file.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("No file to operate on")

    filename = sys.argv[1]
    sheet = openpyxl.load_workbook(filename).active

    sections = []

    parse_sections(sheet, sections)

    for i in range(0, len(sections)):
        sections[i].print_info()

    print("Verify the information above.")

    while 1:
        print("Create .ics? [Y/n]")
        confirm_string = input()
        if confirm_string == "N" or confirm_string == "n":
            print("Cancelled.")
            break
        elif (confirm_string == "Y" or confirm_string == "y"
              or confirm_string == "" or confirm_string.isspace()):
            gen_ics(sections)
            print("Done.")
            print("Output written to out.ics.")
            break
        else:
            print("Invalid input. Try again")
